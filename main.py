import sys
import smtplib
import random
import time

from bs4 import BeautifulSoup
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl import load_workbook

	
def sendEmailToRecipient(server,user,recipient,html) :		
	mail = createEmailMessageFromHTML(user,recipient.strip(),html)
	try:
		server.sendmail(user, recipient.strip(), mail.as_string())
		return True
	except smtplib.SMTPException:
		return False
			

def createEmailMessageFromHTML(user,recipient,html) :
	mail = MIMEMultipart()
	mail['Subject'] = 'Pesquisa da Fernanda'
	mail['From'] = user
	mail['To'] = recipient
	msg = MIMEText(html, 'html')
	mail.attach(msg)

	return mail
	
def replaceNameInHTMLMessage(html,name) :
	soup = BeautifulSoup(html,'lxml')
	for i in soup.find_all(class_='greeting') :
		i.string = "Olá, " + name
	return str(soup)	

	
def authenticateOnGmail(usr,pwd) :
	server = smtplib.SMTP("smtp.gmail.com", 587)
	server.ehlo()
	server.starttls()
	server.login(usr,pwd)
	return server

def getResponseByGenderCount(ws) :
	m = ws['A3'].value
	f = ws['B3'].value

	return {'m':m , 'f':f}
	
def getRecipientCount(m_ws,f_ws) :
	m = m_ws.max_row-1
	f = f_ws.max_row-1
	
	return {'m':m, 'f':f}


def getAvailableRecipients(ws) :
	# Get recipients that havent received an email yet
	available_recipients = {}
	
	for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=3) :
		if row[2].value == 0 :
			available_recipients[row[0].value.lower().title().strip()] = row[1].value.lower().strip()
			
	return available_recipients

def getAvailableRecipientCount(m_d,f_d) :
	m = len(m_d)
	f = len(f_d)
	return {'m':m, 'f':f}

def drawRandomRecipients(recipients,num) :
	#Select random recipient from available
	selected_recipients = {}
	
	for _ in range(0,int(num)) :
		r = random.choice(list(recipients))
		selected_recipients[r] = recipients[r]
		del recipients[r]
		
	return selected_recipients

def updateRecipientTable(fn,wb,ws,r) :
	# Marks on excel table the selected recipients
	for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=3) :
		if row[0].value == r :
			row[2].value = 1
			
	wb.save(fn)

def query_yes_no(msg) :
	yes = {'yes','y','s','sim'}
	no = {'no','n','nao'}
	
	while True :
		choice = input(msg + " s/n ? ").lower()
		
		if choice in yes :
			return True
		if choice in no :
			return False
		else :
			print("Escreva S ou N")

def main(args):
	print("Iniciando...")
	
	# Parse login info
	user = str(args[1])
	password = str(args[2])
	
	
	# Load recipients file to memory
	print("Carregando arquivo...")
	filename = 'emails.xlsx'
	wb = load_workbook(filename)
	geral_sheet = wb['GERAL']
	male_recipients_sheet = wb['HOMENS']
	female_recipients_sheet = wb['MULHERES']
	
	# Load message text to memory
	textfile = 'mensagem.txt'
	with open(textfile, 'rb') as fp:
		textmsg = fp.read()
	
	# Load message html to memory
	htmlfile = 'mensagem.html'
	with open(htmlfile, 'r') as fp:
		htmlmsg = fp.read()
	
	# Get available recipients
	available_male_recipients = getAvailableRecipients(male_recipients_sheet)
	available_female_recipients = getAvailableRecipients(female_recipients_sheet)
	
	# Stats
	num_response = getResponseByGenderCount(geral_sheet)
	num_recipients = getRecipientCount(male_recipients_sheet,female_recipients_sheet)
	num_available_recipients = getAvailableRecipientCount(available_male_recipients,available_female_recipients)
	
	print("Total cadastrados:             {0:4} Homens e {1:4} Mulheres.".format(num_recipients['m'], num_recipients['f']))
	print("Total cadastrados disponíveis: {0:4} Homens e {1:4} Mulheres.".format(num_available_recipients['m'], num_available_recipients['f']))
	print("Total respostas:               {0:4} Homens e {1:4} Mulheres.".format(num_response['m'], num_response['f']))

	# Draw phase
	num_desired_recipients = {}
	
	accepted = False
	while accepted == False :
		num_desired_recipients['m'] = int(input("Digite o numero de e-mails para mandar para homens: "))
		num_desired_recipients['f'] = int(input("Digite o numero de e-mails para mandar para mulheres: "))
		if num_desired_recipients['m']  <= num_available_recipients['m'] and  \
			num_desired_recipients['f'] <= num_available_recipients['f'] :
			accepted = True
		else :
			print("Escolha um número menor.")
			
	# Select female recipients
	accepted = False
	while accepted == False :
		selected_female_recipients = drawRandomRecipients(available_female_recipients.copy(), num_desired_recipients['f'])
		selected_male_recipients = drawRandomRecipients(available_male_recipients.copy(), num_desired_recipients['m'])
	
		print("Homens selecionados:")
		for r in selected_male_recipients :
			print("{0:50} {1:40}".format(r,selected_male_recipients[r]))
		
		print("")
		
		print("Selected female recipients")
		for r in selected_female_recipients :
			print("{0:50} {1:40}".format(r,selected_female_recipients[r]))
		
		accepted = query_yes_no("Você aceita a seleção?")
	
	input("Pressione qualquer tecla para enviar os e-mails aos sorteados.")	
	
	print("Autenticando no servidor de e-mail...")
	server = authenticateOnGmail(user,password)
	
	# Send to female recipients
	print("Enviando para as mulheres...")
	for r in selected_female_recipients :
		name = "Professora " + r.split()[0] + "!"
		m = replaceNameInHTMLMessage(htmlmsg,name)
		if sendEmailToRecipient(server,user,selected_female_recipients[r],m) :
			updateRecipientTable(filename,wb,female_recipients_sheet,r)
			print("E-mail enviado para {}. Tabela atualizada!".format(r))
			time.sleep(5)
		else :
			print("Falha no envio. Você pode estar spammando. Tente de novo mais tarde.")
			break
			
	
	# Send to male recipients
	print("Enviando para os homens...")
	for r in selected_male_recipients :
		name = "Professor " + r.split()[0] + "!"
		m = replaceNameInHTMLMessage(htmlmsg,name)
		if sendEmailToRecipient(server,user,selected_male_recipients[r],m) :
			updateRecipientTable(filename,wb,female_recipients_sheet,r)
			print("E-mail enviado para {}. Tabela atualizada!".format(r))
			time.sleep(5)
		else :
			print("Falha no envio. Você pode estar spammando. Tente de novo mais tarde.")
			break
	
	print("Fechando conexão...")
	server.close()
	
	input("Pronto! Pressione qualquer tecla para fechar.")



if __name__ == '__main__' :
	
	main(sys.argv)
